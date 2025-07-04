import google.generativeai as genai
from openpyxl import load_workbook
import time

# ========== CONFIG ==========
API_KEY = "AIzaSyAfsapZ2Dtwizxr9ggdU387olNydz2mTXM"  # Replace this with your actual Gemini API key
EXCEL_PATH = r"D:\Plants\plants.xlsx"
SAVE_PATH = r"D:\Plants\plants_filled.xlsx"
MODEL_NAME = "models/gemini-1.5-pro-latest"

# ========== INITIALIZE ==========
genai.configure(api_key=API_KEY)
model = genai.GenerativeModel(model_name=MODEL_NAME)

# ========== PROMPT GENERATOR ==========
def generate_prompts(plant_name):
    return {
        "B": f"What is the botanical name of {plant_name}?",
        "C": f"What is the common name of {plant_name}?",
        "D": f"When does {plant_name} bloom?",
        "E": f"What is the maximum height of {plant_name}?",
        "F": f"What light condition does {plant_name} prefer?",
        "G": f"What is the watering schedule of {plant_name}?",
        "H": f"What soil type is suitable for {plant_name}?",
        "I": f"What are the uses of {plant_name}?"
    }

# ========== AI CALL ==========
def ask_gemini(prompt, retries=3):
    for attempt in range(1, retries + 1):
        try:
            response = model.generate_content(prompt)
            return response.text.strip()
        except Exception as e:
            print(f"⚠️ Gemini Error: {e} (Attempt {attempt})")
            time.sleep(5 * attempt)
    return "N/A"

# ========== RESUME FUNCTION ==========
def find_resume_row(ws):
    for row in range(2, 612):
        if not all(ws[f"{col}{row}"].value for col in "BCDEFGHI"):
            return row
    return 612

# ========== MAIN ==========
def run_filler():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    start_row = 58  # Force resume from row 59
    print(f"▶️ Resuming from row {start_row}...")

    for row in range(start_row, 612):
        plant_name = ws[f"A{row}"].value
        if not plant_name:
            continue

        print(f"\n🔄 Row {row} → {plant_name}")
        prompts = generate_prompts(plant_name)

        for col, prompt in prompts.items():
            cell = f"{col}{row}"
            if ws[cell].value:
                continue

            time.sleep(2)  # Prevent hitting limits
            answer = ask_gemini(prompt)

            # Basic filtering
            if "not found" in answer.lower() or "sorry" in answer.lower():
                answer = "Unknown"

            ws[cell] = answer
            print(f"  ➤ {col}: {answer}")

        wb.save(SAVE_PATH)

    print(f"\n✅ Done! File saved to: {SAVE_PATH}")

# ========== EXECUTE ==========
if __name__ == "__main__":
    run_filler()
